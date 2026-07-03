from __future__ import annotations

from collections import defaultdict
from datetime import date
import json
from pathlib import Path
import re
from typing import Any
from xml.etree.ElementTree import iterparse
from zipfile import ZipFile

import pandas as pd
from openpyxl import load_workbook


BACKEND_DIR = Path(__file__).resolve().parents[2]
UPLOAD_DIR = BACKEND_DIR / "uploads"
CACHE_FILE = UPLOAD_DIR / ".dashboard_cache.json"
CACHE_VERSION = 3
MULTI_VALUE_SEPARATOR = "|||"

MONTH_RE = re.compile(r"^(?P<year>20\d{2})[/-](?P<month>\d{1,2})")
COMPANY_HINTS = ("corporation", "company", "manufacturer")
BRAND_HINTS = ("umbrella brand", "brand")
SKU_HINTS = ("sku", "product")
REGION_HINTS = ("region", "territory")
MOLECULE_HINTS = ("molecule",)
MARKET_HINTS = ("rx/otc", "segment", "market")
ATC1_HINTS = ("atc1",)
ATC2_HINTS = ("atc2",)
ATC3_HINTS = ("atc3",)
ATC4_HINTS = ("atc4",)
VALUE_HINTS = ("price", "value", "sales")
UNIT_HINTS = ("unit", "qty", "quantity")


def get_latest_upload() -> Path | None:
    files = [
        path
        for pattern in ("*.xlsx", "*.xls", "*.csv")
        for path in UPLOAD_DIR.glob(pattern)
        if path.is_file()
    ]
    if not files:
        return None
    return max(files, key=lambda path: path.stat().st_mtime)


def get_dashboard_data(
    file_path: str | Path | None = None,
    filters: dict[str, str | None] | None = None,
) -> dict[str, Any]:
    path = Path(file_path) if file_path else get_latest_upload()
    if not path or not path.exists():
        return _empty_dashboard("No uploaded data file found")

    active_filters = _normalize_filters(filters)
    cached = _read_cache(path)
    if cached:
        if cached.get("records"):
            return _build_dashboard_from_records(
                path=path,
                latest_period=cached.get("latest_period"),
                periods=cached.get("periods", []),
                records=cached["records"],
                filters=active_filters,
            )

    try:
        if path.suffix.lower() == ".csv":
            data = _from_flat_dataframe(pd.read_csv(path), path)
        else:
            data = _from_iqvia_workbook(path)
    except Exception:
        try:
            data = _from_flat_dataframe(pd.read_excel(path), path)
        except Exception as exc:
            data = _empty_dashboard(f"Could not read {path.name}: {exc}")
            data["metadata"]["filename"] = path.name

    cache_payload = data.pop("_cache", None)
    if cache_payload:
        _write_cache(path, cache_payload)
        return _build_dashboard_from_records(
            path=path,
            latest_period=cache_payload.get("latest_period"),
            periods=cache_payload.get("periods", []),
            records=cache_payload["records"],
            filters=active_filters,
        )

    _write_cache(path, data)
    return data


def _read_cache(path: Path) -> dict[str, Any] | None:
    try:
        if not CACHE_FILE.exists():
            return None
        payload = json.loads(CACHE_FILE.read_text(encoding="utf-8"))
        if (
            payload.get("source") != str(path)
            or payload.get("mtime") != path.stat().st_mtime
            or payload.get("version") != CACHE_VERSION
        ):
            return None
        return payload.get("data")
    except Exception:
        return None


def _write_cache(path: Path, data: dict[str, Any]) -> None:
    try:
        CACHE_FILE.write_text(
            json.dumps(
                {"source": str(path), "mtime": path.stat().st_mtime, "version": CACHE_VERSION, "data": data},
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
    except Exception:
        pass


def _from_iqvia_workbook(path: Path) -> dict[str, Any]:
    if path.suffix.lower() == ".xlsx":
        return _from_iqvia_xlsx_fast(path)

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active

    header_row_number, header, metric = _find_iqvia_header(worksheet)
    if not header_row_number:
        return _from_flat_dataframe(pd.read_excel(path), path)

    company_idx = _find_column(header, COMPANY_HINTS)
    brand_idx = _find_column(header, BRAND_HINTS)
    sku_idx = _find_column(header, SKU_HINTS)
    region_idx = _find_column(header, REGION_HINTS)

    value_columns: list[dict[str, Any]] = []
    unit_columns: list[dict[str, Any]] = []

    for index, raw_period in enumerate(header):
        period = _parse_period(raw_period)
        if not period:
            continue

        metric_name = _clean(metric[index] if index < len(metric) else "")
        if _has_any(metric_name, VALUE_HINTS):
            value_columns.append({"index": index, "period": period})
        elif _has_any(metric_name, UNIT_HINTS):
            unit_columns.append({"index": index, "period": period})

    value_by_period: dict[date, float] = defaultdict(float)
    units_by_period: dict[date, float] = defaultdict(float)
    company_value: dict[str, float] = defaultdict(float)
    company_units: dict[str, float] = defaultdict(float)
    brand_value: dict[tuple[str, str], float] = defaultdict(float)
    brand_units: dict[tuple[str, str], float] = defaultdict(float)
    sku_value: dict[tuple[str, str, str], float] = defaultdict(float)
    sku_units: dict[tuple[str, str, str], float] = defaultdict(float)
    region_value: dict[str, float] = defaultdict(float)
    row_count = 0

    latest_period = max([column["period"] for column in value_columns], default=None)
    latest_value_indices = [column["index"] for column in value_columns if column["period"] == latest_period]
    latest_unit_indices = [column["index"] for column in unit_columns if column["period"] == latest_period]
    data_start_row = header_row_number + 3
    has_period_totals = False

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=header_row_number + 2, values_only=True),
        start=header_row_number + 2,
    ):
        company = _cell(row, company_idx)
        if company:
            data_start_row = row_number
            break

        row_text = " ".join(str(value) for value in row if value is not None).lower()
        if not has_period_totals and ("total market" in row_text or "total:" in row_text):
            has_period_totals = True
            for column in value_columns:
                value_by_period[column["period"]] += _number(row[column["index"]])
            for column in unit_columns:
                units_by_period[column["period"]] += _number(row[column["index"]])

    for row in worksheet.iter_rows(min_row=data_start_row, values_only=True):
        company = _cell(row, company_idx)
        if not company:
            continue

        company_lower = company.lower()
        if company_lower.startswith("total") or company_lower == "total market:":
            continue

        row_count += 1
        brand = _cell(row, brand_idx) or _cell(row, sku_idx) or "Unknown"
        region = _cell(row, region_idx) or "Unknown"

        latest_value = sum(_number(row[index]) for index in latest_value_indices)
        latest_units = sum(_number(row[index]) for index in latest_unit_indices)
        if latest_period and not has_period_totals:
            value_by_period[latest_period] += latest_value
        if latest_period and not has_period_totals:
            units_by_period[latest_period] += latest_units

        if latest_value:
            company_value[company] += latest_value
            brand_value[(brand, company)] += latest_value
            region_value[region] += latest_value

    total_market = value_by_period.get(latest_period, 0.0) if latest_period else 0.0
    total_units = units_by_period.get(latest_period, 0.0) if latest_period else 0.0
    nobel_sales = sum(value for name, value in company_value.items() if "nobel" in name.lower())
    market_share = (nobel_sales / total_market * 100) if total_market else 0.0

    trend = [
        {
            "month": period.strftime("%b %Y"),
            "sales": round(value_by_period[period] / 1_000_000, 2),
            "units": round(units_by_period.get(period, 0.0)),
        }
        for period in sorted(value_by_period)
    ]

    top_companies = _top_items(company_value, total_market, "name")
    top_brands = [
        {
            "brand": brand,
            "company": company,
            "sales": round(value, 2),
            "share": round(value / total_market * 100, 2) if total_market else 0.0,
        }
        for (brand, company), value in sorted(brand_value.items(), key=lambda item: item[1], reverse=True)[:10]
    ]

    return {
        "kpis": {
            "total_market_value": round(total_market, 2),
            "total_units": round(total_units),
            "nobel_sales": round(nobel_sales, 2),
            "market_share": round(market_share, 2),
            "growth": _growth(value_by_period, latest_period),
        },
        "charts": {
            "sales_trend": trend,
            "market_share": top_companies,
            "regions": _top_items(region_value, total_market, "name"),
        },
        "top_brands": top_brands,
        "metadata": {
            "filename": path.name,
            "rows": row_count,
            "latest_period": latest_period.strftime("%Y-%m") if latest_period else None,
            "message": "Dashboard data calculated from the latest uploaded file",
        },
    }


def _from_iqvia_xlsx(path: Path) -> dict[str, Any]:
    header_row_number = 0
    header: list[Any] = []
    metric: list[Any] = []
    company_idx = brand_idx = sku_idx = region_idx = None
    value_columns: list[dict[str, Any]] = []
    unit_columns: list[dict[str, Any]] = []
    value_by_period: dict[date, float] = defaultdict(float)
    units_by_period: dict[date, float] = defaultdict(float)
    company_value: dict[str, float] = defaultdict(float)
    company_units: dict[str, float] = defaultdict(float)
    brand_value: dict[tuple[str, str], float] = defaultdict(float)
    brand_units: dict[tuple[str, str], float] = defaultdict(float)
    sku_value: dict[tuple[str, str, str], float] = defaultdict(float)
    sku_units: dict[tuple[str, str, str], float] = defaultdict(float)
    region_value: dict[str, float] = defaultdict(float)
    row_count = 0
    latest_period = None
    latest_value_indices: list[int] = []
    latest_unit_indices: list[int] = []
    has_period_totals = False
    data_started = False

    rows = _iter_xlsx_rows(path)

    for row_number, row in rows:
        if not header:
            cleaned = [_clean(value) for value in row]
            if "corporation" in cleaned and any(_parse_period(value) for value in row):
                header_row_number = row_number
                header = row
            continue

        if not metric:
            metric = row
            company_idx = _find_column(tuple(header), COMPANY_HINTS)
            brand_idx = _find_column(tuple(header), BRAND_HINTS)
            sku_idx = _find_column(tuple(header), SKU_HINTS)
            region_idx = _find_column(tuple(header), REGION_HINTS)

            for index, raw_period in enumerate(header):
                period = _parse_period(raw_period)
                if not period:
                    continue

                metric_name = _clean(metric[index] if index < len(metric) else "")
                if _has_any(metric_name, VALUE_HINTS):
                    value_columns.append({"index": index, "period": period})
                elif _has_any(metric_name, UNIT_HINTS):
                    unit_columns.append({"index": index, "period": period})

            latest_period = max([column["period"] for column in value_columns], default=None)
            latest_value_indices = [column["index"] for column in value_columns if column["period"] == latest_period]
            latest_unit_indices = [column["index"] for column in unit_columns if column["period"] == latest_period]
            continue

        if row_number <= header_row_number + 1:
            continue

        company = _cell(tuple(row), company_idx)

        if not data_started:
            if company:
                data_started = True
            else:
                row_text = " ".join(str(value) for value in row if value is not None).lower()
                if not has_period_totals and ("total market" in row_text or "total:" in row_text):
                    has_period_totals = True
                    for column in value_columns:
                        value_by_period[column["period"]] += _number(_row_value(row, column["index"]))
                    for column in unit_columns:
                        units_by_period[column["period"]] += _number(_row_value(row, column["index"]))
                continue

        if not company:
            continue

        company_lower = company.lower()
        if company_lower.startswith("total") or company_lower == "total market:":
            continue

        row_count += 1
        brand = _cell(tuple(row), brand_idx) or _cell(tuple(row), sku_idx) or "Unknown"
        region = _cell(tuple(row), region_idx) or "Unknown"
        latest_value = sum(_number(_row_value(row, index)) for index in latest_value_indices)
        latest_units = sum(_number(_row_value(row, index)) for index in latest_unit_indices)

        if latest_period and not has_period_totals:
            value_by_period[latest_period] += latest_value
            units_by_period[latest_period] += latest_units

        if latest_value:
            company_value[company] += latest_value
            brand_value[(brand, company)] += latest_value
            region_value[region] += latest_value

    if not header or not metric:
        return _from_flat_dataframe(pd.read_excel(path), path)

    total_market = value_by_period.get(latest_period, 0.0) if latest_period else 0.0
    total_units = units_by_period.get(latest_period, 0.0) if latest_period else 0.0
    nobel_sales = sum(value for name, value in company_value.items() if "nobel" in name.lower())
    market_share = (nobel_sales / total_market * 100) if total_market else 0.0

    trend = [
        {
            "month": period.strftime("%b %Y"),
            "sales": round(value_by_period[period] / 1_000_000, 2),
            "units": round(units_by_period.get(period, 0.0)),
        }
        for period in sorted(value_by_period)
    ]

    top_companies = _top_items(company_value, total_market, "name")
    top_brands = [
        {
            "brand": brand,
            "company": company,
            "sales": round(value, 2),
            "share": round(value / total_market * 100, 2) if total_market else 0.0,
        }
        for (brand, company), value in sorted(brand_value.items(), key=lambda item: item[1], reverse=True)[:10]
    ]

    return {
        "kpis": {
            "total_market_value": round(total_market, 2),
            "total_units": round(total_units),
            "nobel_sales": round(nobel_sales, 2),
            "market_share": round(market_share, 2),
            "growth": _growth(value_by_period, latest_period),
        },
        "charts": {
            "sales_trend": trend,
            "market_share": top_companies,
            "regions": _top_items(region_value, total_market, "name"),
        },
        "top_brands": top_brands,
        "metadata": {
            "filename": path.name,
            "rows": row_count,
            "latest_period": latest_period.strftime("%Y-%m") if latest_period else None,
            "message": "Dashboard data calculated from the latest uploaded file",
        },
    }


def _from_iqvia_xlsx_fast(path: Path) -> dict[str, Any]:
    layout = _read_iqvia_xlsx_layout(path)
    if not layout:
        return _from_flat_dataframe(pd.read_excel(path), path)

    header = layout["header"]
    metric = layout["metric"]
    company_idx = _find_column(tuple(header), COMPANY_HINTS)
    brand_idx = _find_column(tuple(header), BRAND_HINTS)
    sku_idx = _find_column(tuple(header), SKU_HINTS)
    region_idx = _find_column(tuple(header), REGION_HINTS)
    molecule_idx = _find_column(tuple(header), MOLECULE_HINTS)
    market_idx = _find_column(tuple(header), MARKET_HINTS)
    atc1_idx = _find_column(tuple(header), ATC1_HINTS)
    atc2_idx = _find_column(tuple(header), ATC2_HINTS)
    atc3_idx = _find_column(tuple(header), ATC3_HINTS)
    atc4_idx = _find_column(tuple(header), ATC4_HINTS)

    value_columns: list[dict[str, Any]] = []
    unit_columns: list[dict[str, Any]] = []
    for index, raw_period in enumerate(header):
        period = _parse_period(raw_period)
        if not period:
            continue

        metric_name = _clean(metric[index] if index < len(metric) else "")
        if _has_any(metric_name, VALUE_HINTS):
            value_columns.append({"index": index, "period": period})
        elif _has_any(metric_name, UNIT_HINTS):
            unit_columns.append({"index": index, "period": period})

    latest_period = max([column["period"] for column in value_columns], default=None)
    latest_value_indices = [column["index"] for column in value_columns if column["period"] == latest_period]
    latest_unit_indices = [column["index"] for column in unit_columns if column["period"] == latest_period]

    wanted_columns = {
        index
        for index in [
            company_idx,
            brand_idx,
            sku_idx,
            region_idx,
            molecule_idx,
            market_idx,
            atc1_idx,
            atc2_idx,
            atc3_idx,
            atc4_idx,
            *[column["index"] for column in value_columns],
            *[column["index"] for column in unit_columns],
        ]
        if index is not None
    }

    value_by_period = layout["value_by_period"]
    units_by_period = layout["units_by_period"]
    has_period_totals = bool(value_by_period)
    company_value: dict[str, float] = defaultdict(float)
    brand_value: dict[tuple[str, str], float] = defaultdict(float)
    brand_units: dict[tuple[str, str], float] = defaultdict(float)
    region_value: dict[str, float] = defaultdict(float)
    records: list[dict[str, Any]] = []
    row_count = 0

    for _, row in _iter_xlsx_rows(path, wanted_columns=wanted_columns, min_row=layout["data_start_row"]):
        company = _cell(tuple(row), company_idx)
        if not company:
            continue

        company_lower = company.lower()
        if company_lower.startswith("total") or company_lower == "total market:":
            continue

        row_count += 1
        brand = _cell(tuple(row), brand_idx) or _cell(tuple(row), sku_idx) or "Unknown"
        sku = _cell(tuple(row), sku_idx) or brand
        region = _cell(tuple(row), region_idx) or "Unknown"
        molecule = _cell(tuple(row), molecule_idx) or "Unknown"
        market = _cell(tuple(row), market_idx) or "Unknown"
        atc1 = _cell(tuple(row), atc1_idx) or "Unknown"
        atc2 = _cell(tuple(row), atc2_idx) or "Unknown"
        atc3 = _cell(tuple(row), atc3_idx) or "Unknown"
        atc4 = _cell(tuple(row), atc4_idx) or "Unknown"
        latest_value = sum(_number(_row_value(row, index)) for index in latest_value_indices)
        latest_units = sum(_number(_row_value(row, index)) for index in latest_unit_indices)
        values_by_period = {
            column["period"].strftime("%Y-%m"): round(_number(_row_value(row, column["index"])), 2)
            for column in value_columns
            if _number(_row_value(row, column["index"]))
        }
        units_by_period_for_row = {
            column["period"].strftime("%Y-%m"): round(_number(_row_value(row, column["index"])))
            for column in unit_columns
            if _number(_row_value(row, column["index"]))
        }

        if latest_period and not has_period_totals:
            value_by_period[latest_period] += latest_value
            units_by_period[latest_period] += latest_units

        if latest_value:
            company_value[company] += latest_value
            brand_value[(brand, company)] += latest_value
            region_value[region] += latest_value
            records.append(
                {
                    "company": company,
                    "brand": brand,
                    "sku": sku,
                    "region": region,
                    "molecule": molecule,
                    "market": market,
                    "atc1": atc1,
                    "atc2": atc2,
                    "atc3": atc3,
                    "atc4": atc4,
                    "value": round(latest_value, 2),
                    "units": round(latest_units),
                    "values_by_period": values_by_period,
                    "units_by_period": units_by_period_for_row,
                }
            )

    data = _build_dashboard_payload(
        path=path,
        latest_period=latest_period,
        value_by_period=value_by_period,
        units_by_period=units_by_period,
        company_value=company_value,
        brand_value=brand_value,
        region_value=region_value,
        row_count=row_count,
    )
    data["_cache"] = {
        "latest_period": latest_period.strftime("%Y-%m") if latest_period else None,
        "periods": sorted({column["period"].strftime("%Y-%m") for column in value_columns}),
        "records": records,
    }
    return data


def _read_iqvia_xlsx_layout(path: Path) -> dict[str, Any] | None:
    header: list[Any] = []
    metric: list[Any] = []
    value_by_period: dict[date, float] = defaultdict(float)
    units_by_period: dict[date, float] = defaultdict(float)
    header_row_number = 0
    company_idx = None
    value_columns: list[dict[str, Any]] = []
    unit_columns: list[dict[str, Any]] = []
    has_period_totals = False

    for row_number, row in _iter_xlsx_rows(path):
        if not header:
            cleaned = [_clean(value) for value in row]
            if "corporation" in cleaned and any(_parse_period(value) for value in row):
                header = row
                header_row_number = row_number
            continue

        if not metric:
            metric = row
            company_idx = _find_column(tuple(header), COMPANY_HINTS)
            for index, raw_period in enumerate(header):
                period = _parse_period(raw_period)
                if not period:
                    continue
                metric_name = _clean(metric[index] if index < len(metric) else "")
                if _has_any(metric_name, VALUE_HINTS):
                    value_columns.append({"index": index, "period": period})
                elif _has_any(metric_name, UNIT_HINTS):
                    unit_columns.append({"index": index, "period": period})
            continue

        company = _cell(tuple(row), company_idx)
        if company:
            return {
                "header": header,
                "metric": metric,
                "header_row_number": header_row_number,
                "data_start_row": row_number,
                "value_by_period": value_by_period,
                "units_by_period": units_by_period,
            }

        row_text = " ".join(str(value) for value in row if value is not None).lower()
        if not has_period_totals and ("total market" in row_text or "total:" in row_text):
            has_period_totals = True
            for column in value_columns:
                value_by_period[column["period"]] += _number(_row_value(row, column["index"]))
            for column in unit_columns:
                units_by_period[column["period"]] += _number(_row_value(row, column["index"]))

    return None


def _build_dashboard_payload(
    path: Path,
    latest_period: date | None,
    value_by_period: dict[date, float],
    units_by_period: dict[date, float],
    company_value: dict[str, float],
    brand_value: dict[tuple[str, str], float],
    region_value: dict[str, float],
    row_count: int,
) -> dict[str, Any]:
    total_market = value_by_period.get(latest_period, 0.0) if latest_period else 0.0
    total_units = units_by_period.get(latest_period, 0.0) if latest_period else 0.0
    nobel_sales = sum(value for name, value in company_value.items() if "nobel" in name.lower())
    market_share = (nobel_sales / total_market * 100) if total_market else 0.0

    trend = [
        {
            "month": period.strftime("%b %Y"),
            "sales": round(value_by_period[period] / 1_000_000, 2),
            "units": round(units_by_period.get(period, 0.0)),
        }
        for period in sorted(value_by_period)
    ]

    return {
        "kpis": {
            "total_market_value": round(total_market, 2),
            "total_units": round(total_units),
            "nobel_sales": round(nobel_sales, 2),
            "market_share": round(market_share, 2),
            "growth": _growth(value_by_period, latest_period),
        },
        "charts": {
            "sales_trend": trend,
            "market_share": _top_items(company_value, total_market, "name"),
            "regions": _top_items(region_value, total_market, "name"),
        },
        "top_brands": [
            {
                "brand": brand,
                "company": company,
                "sales": round(value, 2),
                "share": round(value / total_market * 100, 2) if total_market else 0.0,
            }
            for (brand, company), value in sorted(brand_value.items(), key=lambda item: item[1], reverse=True)[:10]
        ],
        "metadata": {
            "filename": path.name,
            "rows": row_count,
            "latest_period": latest_period.strftime("%Y-%m") if latest_period else None,
            "message": "Dashboard data calculated from the latest uploaded file",
        },
    }


def _build_dashboard_from_records(
    path: Path,
    latest_period: str | None,
    periods: list[str],
    records: list[dict[str, Any]],
    filters: dict[str, str],
) -> dict[str, Any]:
    dimension_filters = {key: value for key, value in filters.items() if key in FILTER_KEYS}
    selected_periods = _selected_periods(periods, latest_period, filters)
    previous_periods = _previous_periods(periods, selected_periods)
    filtered = [record for record in records if _record_matches(record, dimension_filters)]

    company_value: dict[str, float] = defaultdict(float)
    company_units: dict[str, float] = defaultdict(float)
    brand_value: dict[tuple[str, str], float] = defaultdict(float)
    brand_units: dict[tuple[str, str], float] = defaultdict(float)
    sku_value: dict[tuple[str, str, str], float] = defaultdict(float)
    sku_units: dict[tuple[str, str, str], float] = defaultdict(float)
    region_value: dict[str, float] = defaultdict(float)
    trend_value: dict[str, float] = defaultdict(float)
    trend_units: dict[str, float] = defaultdict(float)
    total_market = 0.0
    total_units = 0.0
    previous_market = 0.0

    for record in filtered:
        values_by_period = record.get("values_by_period") or {}
        units_by_period = record.get("units_by_period") or {}
        value = sum(_number(values_by_period.get(period)) for period in selected_periods)
        units = sum(_number(units_by_period.get(period)) for period in selected_periods)
        previous_market += sum(_number(values_by_period.get(period)) for period in previous_periods)
        if not value and not units:
            continue

        company = str(record.get("company") or "Unknown")
        brand = str(record.get("brand") or "Unknown")
        sku = str(record.get("sku") or "Unknown")
        region = str(record.get("region") or "Unknown")
        total_market += value
        total_units += units
        company_value[company] += value
        company_units[company] += units
        brand_value[(brand, company)] += value
        brand_units[(brand, company)] += units
        sku_value[(sku, brand, company)] += value
        sku_units[(sku, brand, company)] += units
        region_value[region] += value

        for period in periods:
            period_value = _number(values_by_period.get(period))
            period_units = _number(units_by_period.get(period))
            if period_value or period_units:
                trend_value[period] += period_value
                trend_units[period] += period_units

    top_company_value = max(company_value.values(), default=0.0)
    market_share = (top_company_value / total_market * 100) if total_market else 0.0
    growth = ((total_market - previous_market) / previous_market * 100) if previous_market else 0.0
    trend = [
        {
            "month": _period_label(period),
            "sales": round(trend_value[period] / 1_000_000, 2),
            "units": round(trend_units[period]),
        }
        for period in periods
        if trend_value.get(period) or trend_units.get(period)
    ]
    cagr = _calculate_cagr(trend_value)

    return {
        "kpis": {
            "total_market_value": round(total_market, 2),
            "total_units": round(total_units),
            "top_company_sales": round(top_company_value, 2),
            "market_share": round(market_share, 2),
            "growth": round(growth, 2),
            "evolution_index": round(total_market / previous_market * 100, 2) if previous_market else 0.0,
            "cagr": cagr,
        },
        "charts": {
            "sales_trend": trend,
            "market_share": _top_items(company_value, total_market, "name"),
            "regions": _top_items(region_value, total_market, "name"),
        },
        "top_brands": [
            {
                "brand": brand,
                "company": company,
                "sales": round(value, 2),
                "units": round(brand_units[(brand, company)]),
                "share": round(value / total_market * 100, 2) if total_market else 0.0,
            }
            for (brand, company), value in sorted(brand_value.items(), key=lambda item: item[1], reverse=True)[:10]
        ],
        "top_companies": [
            {
                "company": company,
                "sales": round(value, 2),
                "units": round(company_units[company]),
                "share": round(value / total_market * 100, 2) if total_market else 0.0,
            }
            for company, value in sorted(company_value.items(), key=lambda item: item[1], reverse=True)[:10]
        ],
        "top_skus": [
            {
                "sku": sku,
                "brand": brand,
                "company": company,
                "sales": round(value, 2),
                "units": round(sku_units[(sku, brand, company)]),
                "share": round(value / total_market * 100, 2) if total_market else 0.0,
            }
            for (sku, brand, company), value in sorted(sku_value.items(), key=lambda item: item[1], reverse=True)[:20]
        ],
        "filter_options": _build_filter_options(records, dimension_filters),
        "period_options": _build_period_options(periods, latest_period),
        "metadata": {
            "filename": path.name,
            "rows": len(filtered),
            "total_rows": len(records),
            "latest_period": latest_period,
            "selected_periods": selected_periods,
            "message": "Dashboard data calculated from the latest uploaded file",
            "filters": filters,
        },
    }


FILTER_KEYS = {"company", "brand", "region", "molecule", "sku", "market", "atc1", "atc2", "atc3", "atc4"}
PERIOD_KEYS = {"period_type", "year", "selected_years", "month", "selected_months", "quarter"}


def _build_filter_options(records: list[dict[str, Any]], filters: dict[str, str]) -> dict[str, list[str]]:
    return {
        key: _unique_options([record for record in records if _record_matches(record, _filters_except(filters, key))], key)
        for key in ("company", "brand", "region", "molecule", "sku", "market", "atc1", "atc2", "atc3", "atc4")
    }


def _unique_options(records: list[dict[str, Any]], key: str, limit: int = 250) -> list[str]:
    totals: dict[str, float] = defaultdict(float)
    for record in records:
        value = str(record.get(key) or "").strip()
        if value and value.lower() != "unknown":
            values_by_period = record.get("values_by_period") or {}
            totals[value] += sum(_number(amount) for amount in values_by_period.values()) or _number(record.get("value"))
    return [name for name, _ in sorted(totals.items(), key=lambda item: item[1], reverse=True)[:limit]]


def _record_matches(record: dict[str, Any], filters: dict[str, str]) -> bool:
    for key, expected in filters.items():
        actual = str(record.get(key) or "")
        expected_values = _split_filter_values(expected)
        if expected_values and actual.lower() not in expected_values:
            return False
    return True


def _split_filter_values(value: str) -> set[str]:
    return {
        part.strip().lower()
        for part in value.split(MULTI_VALUE_SEPARATOR)
        if part.strip()
    }


def _normalize_filters(filters: dict[str, str | None] | None) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for key, value in (filters or {}).items():
        if key not in FILTER_KEYS | PERIOD_KEYS:
            continue
        clean_value = str(value or "").strip()
        if clean_value and clean_value.lower() not in {"all", "all companies", "all brands", "all regions"}:
            normalized[key] = clean_value
    return normalized


def _filters_except(filters: dict[str, str], excluded_key: str) -> dict[str, str]:
    return {key: value for key, value in filters.items() if key != excluded_key}


def _selected_periods(periods: list[str], latest_period: str | None, filters: dict[str, str]) -> list[str]:
    if not periods:
        return []

    selected_type = filters.get("period_type", "MONTH").upper()
    if selected_type == "MTH":
        selected_type = "MONTH"
    selected_year = int(filters.get("year") or (latest_period or periods[-1])[:4])
    selected_years = [
        int(year)
        for year in (filters.get("selected_years") or str(selected_year)).split(",")
        if year.strip().isdigit()
    ]
    selected_month = int(filters.get("month") or (latest_period or periods[-1])[5:7])
    selected_quarter = int(str(filters.get("quarter") or ((selected_month - 1) // 3 + 1)).replace("Q", ""))

    if selected_type == "MONTH":
        targets = {f"{year}-{selected_month:02d}" for year in selected_years}
        return [period for period in periods if period in targets]

    if selected_type == "MONTHS":
        selected_months = [
            int(month)
            for month in filters.get("selected_months", "").split(",")
            if month.strip().isdigit()
        ]
        targets = {f"{year}-{month:02d}" for year in selected_years for month in selected_months}
        return [period for period in periods if period in targets]

    if selected_type == "YTD":
        return [
            period
            for period in periods
            if int(period[:4]) in selected_years and int(period[5:7]) <= selected_month
        ]

    if selected_type == "QTR":
        start_month = (selected_quarter - 1) * 3 + 1
        quarter_months = {f"{year}-{month:02d}" for year in selected_years for month in range(start_month, start_month + 3)}
        return [period for period in periods if period in quarter_months]

    if selected_type == "FULL_YEAR":
        return [period for period in periods if int(period[:4]) in selected_years]

    if selected_type == "MAT":
        end = f"{selected_year}-{selected_month:02d}"
        eligible = [period for period in periods if period <= end]
        return eligible[-12:]

    return [latest_period or periods[-1]]


def _previous_periods(periods: list[str], selected_periods: list[str]) -> list[str]:
    if not selected_periods:
        return []
    first_index = periods.index(selected_periods[0]) if selected_periods[0] in periods else 0
    length = len(selected_periods)
    start = max(first_index - length, 0)
    return periods[start:first_index]


def _calculate_cagr(trend_value: dict[str, float]) -> float:
    populated = [(period, value) for period, value in sorted(trend_value.items()) if value > 0]
    if len(populated) < 2:
        return 0.0
    first_period, first_value = populated[0]
    last_period, last_value = populated[-1]
    first_year, first_month = [int(part) for part in first_period.split("-")]
    last_year, last_month = [int(part) for part in last_period.split("-")]
    months = max((last_year - first_year) * 12 + (last_month - first_month), 1)
    years = months / 12
    if first_value <= 0 or last_value <= 0 or years <= 0:
        return 0.0
    return round(((last_value / first_value) ** (1 / years) - 1) * 100, 2)


def _build_period_options(periods: list[str], latest_period: str | None) -> dict[str, Any]:
    years = sorted({period[:4] for period in periods}, reverse=True)
    latest = latest_period or (periods[-1] if periods else None)
    return {
        "period_types": ["MONTH", "MONTHS", "YTD", "QTR", "MAT", "FULL_YEAR"],
        "years": years,
        "months": [{"value": f"{month:02d}", "label": _month_name(month)} for month in range(1, 13)],
        "quarters": ["Q1", "Q2", "Q3", "Q4"],
        "default": {
            "period_type": "MONTH",
            "year": latest[:4] if latest else "",
            "selected_years": latest[:4] if latest else "",
            "month": latest[5:7] if latest else "",
            "selected_months": latest[5:7] if latest else "",
            "quarter": f"Q{((int(latest[5:7]) - 1) // 3 + 1)}" if latest else "Q1",
        },
    }


def _period_label(period: str) -> str:
    year, month = period.split("-")
    return f"{_month_name(int(month))[:3]} {year}"


def _month_name(month: int) -> str:
    return [
        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July",
        "August",
        "September",
        "October",
        "November",
        "December",
    ][month - 1]


def _from_flat_dataframe(dataframe: pd.DataFrame, path: Path) -> dict[str, Any]:
    columns = {_clean(column): column for column in dataframe.columns}
    value_column = _first_matching(columns, VALUE_HINTS)
    unit_column = _first_matching(columns, UNIT_HINTS)
    company_column = _first_matching(columns, COMPANY_HINTS)
    brand_column = _first_matching(columns, BRAND_HINTS)
    period_column = _first_matching(columns, ("period", "month", "date"))

    total_market = float(pd.to_numeric(dataframe[value_column], errors="coerce").fillna(0).sum()) if value_column else 0.0
    total_units = float(pd.to_numeric(dataframe[unit_column], errors="coerce").fillna(0).sum()) if unit_column else 0.0

    nobel_sales = 0.0
    if company_column and value_column:
        mask = dataframe[company_column].astype(str).str.contains("nobel", case=False, na=False)
        nobel_sales = float(pd.to_numeric(dataframe.loc[mask, value_column], errors="coerce").fillna(0).sum())

    top_brands = []
    if brand_column and value_column:
        grouped = dataframe.groupby(brand_column, dropna=True)[value_column].sum().sort_values(ascending=False).head(10)
        top_brands = [
            {
                "brand": str(brand),
                "company": "",
                "sales": round(float(value), 2),
                "share": round(float(value) / total_market * 100, 2) if total_market else 0.0,
            }
            for brand, value in grouped.items()
        ]

    trend = []
    if period_column and value_column:
        period_frame = dataframe.copy()
        period_frame[period_column] = pd.to_datetime(period_frame[period_column], errors="coerce")
        grouped = period_frame.dropna(subset=[period_column]).groupby(pd.Grouper(key=period_column, freq="MS"))[value_column].sum()
        trend = [{"month": index.strftime("%b %Y"), "sales": round(float(value) / 1_000_000, 2)} for index, value in grouped.items()]

    return {
        "kpis": {
            "total_market_value": round(total_market, 2),
            "total_units": round(total_units),
            "nobel_sales": round(nobel_sales, 2),
            "market_share": round(nobel_sales / total_market * 100, 2) if total_market else 0.0,
            "growth": 0.0,
        },
        "charts": {"sales_trend": trend, "market_share": []},
        "top_brands": top_brands,
        "metadata": {
            "filename": path.name,
            "rows": len(dataframe),
            "latest_period": None,
            "message": "Dashboard data calculated from uploaded table",
        },
    }


def _find_iqvia_header(worksheet) -> tuple[int | None, tuple[Any, ...], tuple[Any, ...]]:
    previous = None
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        cleaned = [_clean(value) for value in row]
        if "corporation" in cleaned and any(_parse_period(value) for value in row):
            metric = next(worksheet.iter_rows(min_row=row_number + 1, max_row=row_number + 1, values_only=True))
            return row_number, row, metric
        previous = row
    return None, previous or (), ()


def _iter_xlsx_rows(
    path: Path,
    wanted_columns: set[int] | None = None,
    min_row: int | None = None,
):
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    shared_strings = _read_shared_strings(path)

    with ZipFile(path) as archive:
        sheet_name = _first_sheet_name(archive)
        with archive.open(sheet_name) as sheet:
            for _, row_element in iterparse(sheet, events=("end",)):
                if row_element.tag != f"{namespace}row":
                    continue

                row_number = int(row_element.attrib.get("r", "0") or 0)
                if min_row and row_number < min_row:
                    row_element.clear()
                    continue

                values: dict[int, Any] = {}
                max_index = -1

                for cell in row_element.findall(f"{namespace}c"):
                    reference = cell.attrib.get("r", "")
                    column_index = _column_index(reference)
                    if column_index < 0:
                        continue
                    if wanted_columns is not None and column_index not in wanted_columns:
                        continue

                    max_index = max(max_index, column_index)
                    values[column_index] = _cell_value(cell, shared_strings, namespace)

                row = [None] * (max_index + 1)
                for index, value in values.items():
                    row[index] = value

                row_element.clear()
                yield row_number, row


def _read_shared_strings(path: Path) -> list[str]:
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    strings: list[str] = []

    with ZipFile(path) as archive:
        if "xl/sharedStrings.xml" not in archive.namelist():
            return strings

        with archive.open("xl/sharedStrings.xml") as shared:
            for _, element in iterparse(shared, events=("end",)):
                if element.tag == f"{namespace}si":
                    texts = [text_element.text or "" for text_element in element.iter(f"{namespace}t")]
                    strings.append("".join(texts))
                    element.clear()

    return strings


def _first_sheet_name(archive: ZipFile) -> str:
    names = archive.namelist()
    for name in names:
        if name.startswith("xl/worksheets/sheet") and name.endswith(".xml"):
            return name
    raise ValueError("No worksheet found in xlsx file")


def _cell_value(cell, shared_strings: list[str], namespace: str) -> Any:
    cell_type = cell.attrib.get("t")

    if cell_type == "inlineStr":
        text = cell.find(f"{namespace}is/{namespace}t")
        return text.text if text is not None else ""

    value = cell.find(f"{namespace}v")
    raw = value.text if value is not None else ""

    if cell_type == "s":
        index = int(raw or 0)
        return shared_strings[index] if index < len(shared_strings) else ""

    if cell_type == "str":
        return raw

    return _number(raw) if raw != "" else None


def _column_index(reference: str) -> int:
    letters = ""
    for char in reference:
        if char.isalpha():
            letters += char.upper()
        else:
            break

    if not letters:
        return -1

    index = 0
    for char in letters:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def _find_column(header: tuple[Any, ...], hints: tuple[str, ...]) -> int | None:
    for index, value in enumerate(header):
        text = _clean(value)
        if text and any(hint in text for hint in hints):
            return index
    return None


def _parse_period(value: Any) -> date | None:
    match = MONTH_RE.match(str(value or "").strip())
    if not match:
        return None
    return date(int(match.group("year")), int(match.group("month")), 1)


def _growth(values: dict[date, float], latest_period: date | None) -> float:
    if not latest_period:
        return 0.0
    periods = sorted(values)
    if latest_period not in periods:
        return 0.0
    latest_index = periods.index(latest_period)
    if latest_index == 0:
        return 0.0
    previous = values[periods[latest_index - 1]]
    if not previous:
        return 0.0
    return round((values[latest_period] - previous) / previous * 100, 2)


def _top_items(values: dict[str, float], total: float, name_key: str) -> list[dict[str, Any]]:
    return [
        {
            name_key: name,
            "value": round(value, 2),
            "share": round(value / total * 100, 2) if total else 0.0,
        }
        for name, value in sorted(values.items(), key=lambda item: item[1], reverse=True)[:10]
    ]


def _first_matching(columns: dict[str, str], hints: tuple[str, ...]) -> str | None:
    for clean_name, original_name in columns.items():
        if any(hint in clean_name for hint in hints):
            return original_name
    return None


def _has_any(value: str, hints: tuple[str, ...]) -> bool:
    return any(hint in value for hint in hints)


def _cell(row: tuple[Any, ...], index: int | None) -> str:
    if index is None or index >= len(row) or row[index] is None:
        return ""
    return str(row[index]).strip()


def _row_value(row: list[Any], index: int) -> Any:
    if index >= len(row):
        return None
    return row[index]


def _number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _clean(value: Any) -> str:
    return str(value or "").strip().lower()


def _empty_dashboard(message: str) -> dict[str, Any]:
    return {
        "kpis": {
            "total_market_value": 0,
            "total_units": 0,
            "nobel_sales": 0,
            "market_share": 0,
            "growth": 0,
        },
        "charts": {"sales_trend": [], "market_share": [], "regions": []},
        "top_brands": [],
        "metadata": {"filename": None, "rows": 0, "latest_period": None, "message": message},
    }
