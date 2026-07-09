from datetime import datetime
from pathlib import Path
import re
from typing import Any

from fastapi import APIRouter, BackgroundTasks, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
import pandas as pd

from app.services.analytics_service import get_dashboard_data
from app.services.file_registry import activate_upload, delete_upload, list_uploads, register_upload
from app.services.performance_cockpit_service import get_performance_cockpit_data


router = APIRouter(
    prefix="/upload",
    tags=["Upload"],
)

UPLOAD_DIR = Path(__file__).resolve().parents[2] / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

DESTINATIONS = {"market_intelligence", "performance_cockpit"}
SOURCES_BY_DESTINATION = {
    "market_intelligence": {"iqvia"},
    "performance_cockpit": {"performance", "secondary_sales"},
}
PERIOD_RE = re.compile(r"(20\d{2})[/-](0?[1-9]|1[0-2])")


@router.get("/")
def get_uploads():
    return {"files": list_uploads()}


@router.post("/")
async def upload_file(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    country: str = Form("Uzbekistan"),
    destination: str = Form(...),
    source: str = Form(...),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="Only Excel or CSV files are supported")
    if destination not in DESTINATIONS:
        raise HTTPException(status_code=400, detail="Please select data destination before uploading.")
    if source not in SOURCES_BY_DESTINATION[destination]:
        raise HTTPException(status_code=400, detail="Invalid file type for selected data destination")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = Path(file.filename).name
    prefix = "market" if destination == "market_intelligence" else "performance"
    file_path = UPLOAD_DIR / f"{prefix}_{timestamp}_{safe_name}"

    with open(file_path, "wb") as buffer:
        buffer.write(await file.read())

    upload_info = register_upload(
        file_path,
        safe_name,
        country,
        destination=destination,
        source=source,
        make_active=True,
    )
    summary = _inspect_upload(file_path)
    if destination == "market_intelligence":
        background_tasks.add_task(_warm_dashboard_cache, file_path)
    if destination == "performance_cockpit":
        background_tasks.add_task(get_performance_cockpit_data, {})

    return {
        "status": "success",
        "filename": file.filename,
        "message": "File uploaded. Analytics will refresh in the background.",
        "path": str(file_path),
        "file": upload_info,
        "dashboard": None,
        "summary": {
            "destination": destination,
            "source": source,
            "file_name": file.filename,
            **summary,
            "status": "uploaded",
        },
    }


@router.post("/{file_id}/activate")
def set_active_upload(file_id: str, background_tasks: BackgroundTasks):
    upload_info = activate_upload(file_id)
    if not upload_info:
        raise HTTPException(status_code=404, detail="File not found")
    if upload_info.get("destination") == "market_intelligence":
        background_tasks.add_task(_warm_dashboard_cache, None)
    if upload_info.get("destination") == "performance_cockpit":
        background_tasks.add_task(get_performance_cockpit_data, {})
    return {
        "status": "success",
        "file": upload_info,
        "dashboard": None,
    }


@router.delete("/{file_id}")
def remove_upload(file_id: str):
    if not delete_upload(file_id):
        raise HTTPException(status_code=404, detail="File not found")
    return {"status": "success", "files": list_uploads()}


def _warm_dashboard_cache(file_path: Path | None = None) -> None:
    try:
        get_dashboard_data(file_path)
    except Exception:
        pass


def _inspect_upload(path: Path) -> dict[str, Any]:
    if path.suffix.lower() == ".csv":
        frame = pd.read_csv(path, nrows=50)
        full_rows = sum(1 for _ in path.open("r", encoding="utf-8", errors="ignore")) - 1
        columns = [str(column) for column in frame.columns]
        return {
            "rows": max(full_rows, 0),
            "columns": len(columns),
            "detected_periods": _periods_from_values(columns + frame.astype(str).head(10).values.flatten().tolist()),
        }

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = max((worksheet.max_row or 1) - 1, 0)
    columns = worksheet.max_column or 0
    values: list[Any] = []
    for row in worksheet.iter_rows(min_row=1, max_row=min(12, worksheet.max_row or 1), values_only=True):
        values.extend(value for value in row if value is not None)
    return {
        "rows": rows,
        "columns": columns,
        "detected_periods": _periods_from_values(values),
    }


def _periods_from_values(values: list[Any]) -> list[str]:
    periods: set[str] = set()
    for value in values:
        for year, month in PERIOD_RE.findall(str(value)):
            periods.add(f"{year}-{int(month):02d}")
    return sorted(periods)
